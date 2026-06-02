"""Restaurant admin endpoints — tenant-scoped, auth required."""
import asyncio
import io
from datetime import datetime, timezone, timedelta

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from db import db
from auth import require_restaurant
from whatsapp import notify_order_status
from models import (
    CategoryIn, ProductIn, CouponIn, BannerIn, RestaurantSettings, StatusUpdate,
    ORDER_STATUSES, clean, new_id, now_iso, is_restaurant_open,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


def rid(user):
    return user["restaurant_id"]


# ---------- restaurant config ----------
@router.get("/restaurant")
async def get_restaurant(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)}, {"_id": 0})
    if not r:
        raise HTTPException(status_code=404, detail="Restaurante não encontrado")
    r["is_open"] = is_restaurant_open(r)
    return r


@router.put("/restaurant")
async def update_restaurant(settings: RestaurantSettings, user=Depends(require_restaurant)):
    updates = {k: v for k, v in settings.model_dump().items() if v is not None}
    updates["updated_at"] = now_iso()
    await db.restaurants.update_one({"id": rid(user)}, {"$set": updates})
    r = await db.restaurants.find_one({"id": rid(user)}, {"_id": 0})
    r["is_open"] = is_restaurant_open(r)
    return r


@router.post("/restaurant/toggle-open")
async def toggle_open(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)})
    new_val = not bool(r.get("is_open_manual", True))
    await db.restaurants.update_one({"id": rid(user)}, {"$set": {"is_open_manual": new_val}})
    return {"is_open_manual": new_val}


@router.get("/restaurant/slug")
async def get_restaurant_slug(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)}, {"slug": 1, "_id": 0})
    if not r:
        raise HTTPException(404, "Restaurante não encontrado")
    return {"slug": r.get("slug", "")}


# ---------- categories ----------
@router.get("/categories")
async def list_categories(user=Depends(require_restaurant)):
    return await db.categories.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(500)


@router.post("/categories")
async def create_category(data: CategoryIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.categories.insert_one(doc)
    return clean(doc)


@router.put("/categories/{cid}")
async def update_category(cid: str, data: CategoryIn, user=Depends(require_restaurant)):
    res = await db.categories.update_one(
        {"id": cid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return await db.categories.find_one({"id": cid}, {"_id": 0})


@router.delete("/categories/{cid}")
async def delete_category(cid: str, user=Depends(require_restaurant)):
    await db.categories.delete_one({"id": cid, "restaurant_id": rid(user)})
    await db.products.delete_many({"category_id": cid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- products ----------
@router.get("/products")
async def list_products(user=Depends(require_restaurant)):
    return await db.products.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(1000)


@router.post("/products")
async def create_product(data: ProductIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.products.insert_one(doc)
    return clean(doc)


@router.put("/products/{pid}")
async def update_product(pid: str, data: ProductIn, user=Depends(require_restaurant)):
    res = await db.products.update_one(
        {"id": pid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return await db.products.find_one({"id": pid}, {"_id": 0})


@router.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_restaurant)):
    await db.products.delete_one({"id": pid, "restaurant_id": rid(user)})
    return {"ok": True}


@router.get("/products/export")
async def export_products(user=Depends(require_restaurant)):
    products = await db.products.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(1000)

    # Build category id -> name map
    cat_ids = list({p.get("category_id") for p in products if p.get("category_id")})
    categories = await db.categories.find({"id": {"$in": cat_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    cat_map = {c["id"]: c["name"] for c in categories}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    headers = [
        "name", "description", "price", "promotional_price",
        "category_name", "is_available", "is_best_seller", "is_featured",
        "track_stock", "stock_quantity", "image_url",
    ]
    ws.append(headers)

    for p in products:
        ws.append([
            p.get("name", ""),
            p.get("description", ""),
            p.get("price", 0),
            p.get("promotional_price", ""),
            cat_map.get(p.get("category_id", ""), ""),
            p.get("is_available", True),
            p.get("is_best_seller", False),
            p.get("is_featured", False),
            p.get("track_stock", False),
            p.get("stock_quantity", ""),
            p.get("image_url", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produtos.xlsx"},
    )


@router.post("/products/import")
async def import_products(file: UploadFile = File(...), user=Depends(require_restaurant)):
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Arquivo inválido: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "updated": 0, "errors": []}

    # Determine column positions from header row
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    required_cols = {"name", "price"}
    missing = required_cols - set(col.keys())
    if missing:
        raise HTTPException(400, f"Colunas obrigatórias ausentes: {missing}")

    restaurant_id = rid(user)
    imported = 0
    updated = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            def _get(field, default=None):
                idx = col.get(field)
                if idx is None:
                    return default
                val = row[idx] if idx < len(row) else None
                return val if val is not None else default

            name = str(_get("name", "")).strip()
            if not name:
                errors.append(f"Linha {row_num}: nome vazio, ignorado")
                continue

            try:
                price = float(_get("price", 0) or 0)
            except (ValueError, TypeError):
                errors.append(f"Linha {row_num}: preço inválido para '{name}'")
                continue

            promo_raw = _get("promotional_price")
            try:
                promotional_price = float(promo_raw) if promo_raw not in (None, "", "none", "null") else None
            except (ValueError, TypeError):
                promotional_price = None

            description = str(_get("description", "") or "")
            category_name = str(_get("category_name", "") or "").strip()
            is_available = bool(_get("is_available", True))
            is_best_seller = bool(_get("is_best_seller", False))

            # Resolve or create category
            category_id = None
            if category_name:
                cat = await db.categories.find_one(
                    {"restaurant_id": restaurant_id, "name": {"$regex": f"^{category_name}$", "$options": "i"}}
                )
                if cat:
                    category_id = cat["id"]
                else:
                    category_id = new_id()
                    await db.categories.insert_one({
                        "id": category_id,
                        "restaurant_id": restaurant_id,
                        "name": category_name,
                        "sort_order": 0,
                        "created_at": now_iso(),
                    })

            payload = {
                "name": name,
                "description": description,
                "price": price,
                "promotional_price": promotional_price,
                "category_id": category_id,
                "is_available": is_available,
                "is_best_seller": is_best_seller,
                "updated_at": now_iso(),
            }

            existing = await db.products.find_one(
                {"restaurant_id": restaurant_id, "name": {"$regex": f"^{name}$", "$options": "i"}}
            )
            if existing:
                await db.products.update_one({"id": existing["id"]}, {"$set": payload})
                updated += 1
            else:
                payload.update({
                    "id": new_id(),
                    "restaurant_id": restaurant_id,
                    "is_featured": False,
                    "track_stock": False,
                    "stock_quantity": None,
                    "image_url": None,
                    "sort_order": 0,
                    "created_at": now_iso(),
                })
                await db.products.insert_one(payload)
                imported += 1

        except Exception as exc:
            errors.append(f"Linha {row_num}: erro inesperado — {exc}")

    return {"imported": imported, "updated": updated, "errors": errors}


# ---------- orders ----------
@router.get("/orders")
async def list_orders(status: str = None, user=Depends(require_restaurant)):
    q = {"restaurant_id": rid(user)}
    if status:
        q["status"] = status
    return await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@router.get("/orders/{oid}")
async def get_order(oid: str, user=Depends(require_restaurant)):
    o = await db.orders.find_one({"id": oid, "restaurant_id": rid(user)}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    return o


@router.put("/orders/{oid}/status")
async def update_order_status(oid: str, data: StatusUpdate, user=Depends(require_restaurant)):
    if data.status not in ORDER_STATUSES:
        raise HTTPException(status_code=400, detail="Status inválido")
    res = await db.orders.update_one(
        {"id": oid, "restaurant_id": rid(user)},
        {"$set": {"status": data.status, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    # Fire-and-forget WhatsApp notification
    asyncio.create_task(notify_order_status(order, data.status))
    return order


# ---------- coupons ----------
@router.get("/coupons")
async def list_coupons(user=Depends(require_restaurant)):
    return await db.coupons.find({"restaurant_id": rid(user)}, {"_id": 0}).to_list(200)


@router.post("/coupons")
async def create_coupon(data: CouponIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc["code"] = doc["code"].upper()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "used_count": 0, "created_at": now_iso()})
    await db.coupons.insert_one(doc)
    return clean(doc)


@router.put("/coupons/{cid}")
async def update_coupon(cid: str, data: CouponIn, user=Depends(require_restaurant)):
    payload = data.model_dump()
    payload["code"] = payload["code"].upper()
    res = await db.coupons.update_one({"id": cid, "restaurant_id": rid(user)}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cupom não encontrado")
    return await db.coupons.find_one({"id": cid}, {"_id": 0})


@router.delete("/coupons/{cid}")
async def delete_coupon(cid: str, user=Depends(require_restaurant)):
    await db.coupons.delete_one({"id": cid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- banners ----------
@router.get("/banners")
async def list_banners(user=Depends(require_restaurant)):
    return await db.banners.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(100)


@router.post("/banners")
async def create_banner(data: BannerIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.banners.insert_one(doc)
    return clean(doc)


@router.put("/banners/{bid}")
async def update_banner(bid: str, data: BannerIn, user=Depends(require_restaurant)):
    res = await db.banners.update_one({"id": bid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Banner não encontrado")
    return await db.banners.find_one({"id": bid}, {"_id": 0})


@router.delete("/banners/{bid}")
async def delete_banner(bid: str, user=Depends(require_restaurant)):
    await db.banners.delete_one({"id": bid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- dashboard & reports ----------
def _today_start_iso():
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    except Exception:
        now = datetime.now()
    return now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()


@router.get("/dashboard")
async def dashboard(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)})
    orders = await db.orders.find({"restaurant_id": rid(user)}, {"_id": 0}).to_list(5000)
    today = _today_start_iso()[:10]
    today_orders = [o for o in orders if (o.get("created_at") or "")[:10] == today and o.get("status") != "cancelled"]
    revenue = sum(o["total"] for o in today_orders)
    in_progress = [o for o in orders if o.get("status") in ("pending", "accepted", "preparing", "ready", "out_for_delivery")]

    # top products
    counter = {}
    for o in orders:
        if o.get("status") == "cancelled":
            continue
        for it in o.get("items", []):
            counter[it["product_name"]] = counter.get(it["product_name"], 0) + it.get("quantity", 1)
    top = sorted(counter.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        "orders_today": len(today_orders),
        "revenue_today": round(revenue, 2),
        "avg_ticket": round(revenue / len(today_orders), 2) if today_orders else 0,
        "in_progress": len(in_progress),
        "is_open": is_restaurant_open(r) if r else False,
        "is_open_manual": r.get("is_open_manual", True) if r else True,
        "top_products": [{"name": n, "qty": q} for n, q in top],
        "recent_orders": sorted(orders, key=lambda o: o.get("created_at", ""), reverse=True)[:8],
    }


@router.get("/reports")
async def reports(period: str = "7d", user=Depends(require_restaurant)):
    days = {"today": 1, "7d": 7, "30d": 30, "90d": 90}.get(period, 7)
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    except Exception:
        now = datetime.now()
    start = (now - timedelta(days=days)).isoformat()
    # Fetch ALL orders (including cancelled) so frontend can compute cancel rate
    all_orders = await db.orders.find(
        {"restaurant_id": rid(user), "created_at": {"$gte": start}},
        {"_id": 0},
    ).to_list(5000)
    return {"orders": all_orders}
